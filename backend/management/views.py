import datetime
from django.utils import timezone
from django.db.models import Sum, Q
from django.shortcuts import get_object_or_404
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

from .models import House, Room, Tenant, Contract, Payment, Expense
from .serializers import HouseSerializer, RoomSerializer, TenantSerializer, ContractSerializer, PaymentSerializer, ExpenseSerializer
from .pagination import OptionalPageNumberPagination

class HouseViewSet(viewsets.ModelViewSet):
    """
    API endpoint for Houses (Nhà trọ)
    """
    queryset = House.objects.all()
    serializer_class = HouseSerializer
    pagination_class = OptionalPageNumberPagination

    def get_queryset(self):
        queryset = House.objects.all()
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | 
                Q(address__icontains=search)
            )
        return queryset


class RoomViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows rooms to be viewed or edited.
    """
    queryset = Room.objects.all()
    serializer_class = RoomSerializer
    pagination_class = OptionalPageNumberPagination

    def get_queryset(self):
        queryset = Room.objects.all()
        # Search by code or name
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(room_code__icontains=search) | 
                Q(room_name__icontains=search)
            )
        # Filter by status
        status_filter = self.request.query_params.get('status', None)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
            
        # Filter by house
        house_id = self.request.query_params.get('house_id', None)
        if house_id:
            queryset = queryset.filter(house_id=house_id)
            
        return queryset

    @action(detail=True, methods=['get'], url_path='latest-utility')
    def latest_utility(self, request, pk=None):
        """
        Get the latest electricity and water index of this room from the most recent invoice,
        along with the building's default utility prices.
        """
        room = self.get_object()
        latest_payment = Payment.objects.filter(room=room).order_by('-month', '-created_at').first()
        
        # Get active contract occupants_count
        active_contract = Contract.objects.filter(room=room, is_active=True).first()
        occupants_count = active_contract.occupants_count if active_contract else 1

        house = room.house

        if latest_payment:
            data = {
                "elec_old": latest_payment.elec_new,
                "water_old": latest_payment.water_new,
                "occupants_count": occupants_count,
                "elec_price": house.elec_price,
                "water_price": house.water_price,
                "service_price": house.service_price,
                "internet_price": house.internet_price
            }
        else:
            data = {
                "elec_old": 0,
                "water_old": 0,
                "occupants_count": occupants_count,
                "elec_price": house.elec_price,
                "water_price": house.water_price,
                "service_price": house.service_price,
                "internet_price": house.internet_price
            }
        return Response(data, status=status.HTTP_200_OK)


class TenantViewSet(viewsets.ModelViewSet):
    """
    API endpoint for Tenants.
    """
    queryset = Tenant.objects.all()
    serializer_class = TenantSerializer
    pagination_class = OptionalPageNumberPagination

    def get_queryset(self):
        queryset = Tenant.objects.all()
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(full_name__icontains=search) | 
                Q(phone__icontains=search) | 
                Q(citizen_id__icontains=search)
            )
        return queryset


class ContractViewSet(viewsets.ModelViewSet):
    """
    API endpoint for Contracts.
    """
    queryset = Contract.objects.all()
    serializer_class = ContractSerializer
    pagination_class = OptionalPageNumberPagination

    def get_queryset(self):
        queryset = Contract.objects.all()
        is_active_filter = self.request.query_params.get('is_active', None)
        if is_active_filter is not None:
            is_active_bool = is_active_filter.lower() in ('true', '1', 't')
            queryset = queryset.filter(is_active=is_active_bool)
            
        house_id = self.request.query_params.get('house_id', None)
        if house_id:
            queryset = queryset.filter(room__house_id=house_id)
            
        return queryset

    @action(detail=True, methods=['post'], url_path='terminate')
    def terminate(self, request, pk=None):
        """
        Custom action to terminate a contract
        """
        contract = self.get_object()
        if not contract.is_active:
            return Response(
                {"detail": "Hợp đồng này đã kết thúc trước đó."}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        contract.is_active = False
        # Set end_date to today if it is in the future
        today = timezone.now().date()
        if contract.end_date > today:
            contract.end_date = today
        
        from django.core.exceptions import ValidationError
        try:
            contract.save()
        except ValidationError as e:
            return Response(
                {"detail": e.messages[0] if hasattr(e, 'messages') else str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        return Response(
            {
                "status": "Hợp đồng đã được kết thúc thành công.",
                "is_active": contract.is_active,
                "end_date": contract.end_date
            }, 
            status=status.HTTP_200_OK
        )


class PaymentViewSet(viewsets.ModelViewSet):
    """
    API endpoint for Payments.
    """
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    pagination_class = OptionalPageNumberPagination

    def get_queryset(self):
        queryset = Payment.objects.all()
        
        # Filter by month (YYYY-MM)
        month = self.request.query_params.get('month', None)
        if month:
            queryset = queryset.filter(month=month)
            
        # Filter by payment status (paid/unpaid)
        payment_status = self.request.query_params.get('status', None)
        if payment_status:
            queryset = queryset.filter(payment_status=payment_status)
            
        # Filter by room
        room_id = self.request.query_params.get('room_id', None)
        if room_id:
            queryset = queryset.filter(room_id=room_id)
            
        # Filter by house
        house_id = self.request.query_params.get('house_id', None)
        if house_id:
            queryset = queryset.filter(room__house_id=house_id)
            
        return queryset

    @action(detail=True, methods=['post'], url_path='confirm')
    def confirm_payment(self, request, pk=None):
        """
        Custom action to confirm a payment invoice
        """
        payment = self.get_object()
        if payment.payment_status == 'paid':
            return Response(
                {"detail": "Hóa đơn này đã được thanh toán rồi."}, 
                status=status.HTTP_400_BAD_REQUEST
            )
            
        payment.payment_status = 'paid'
        payment.save(update_fields=['payment_status'])
        return Response(
            {
                "status": "Đã xác nhận thanh toán hóa đơn.",
                "payment_status": payment.payment_status
            },
            status=status.HTTP_200_OK
        )

    @action(detail=True, methods=['get'], url_path='export-excel')
    def export_excel(self, request, pk=None):
        """
        Custom action to export service fees as a styled Excel sheet matching the user's design.
        """
        payment = self.get_object()
        room = payment.room
        house = room.house
        
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from django.http import HttpResponse
        
        # 1. Create workbook and select active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Phiếu thu phòng {room.room_code}"
        
        # Show gridlines
        ws.sheet_view.showGridLines = True
        
        # Set column widths
        ws.column_dimensions['A'].width = 6   # STT
        ws.column_dimensions['B'].width = 25  # Tên dịch vụ
        ws.column_dimensions['C'].width = 12  # Chỉ số mới
        ws.column_dimensions['D'].width = 12  # Chỉ số cũ
        ws.column_dimensions['E'].width = 12  # Khối mới
        ws.column_dimensions['F'].width = 12  # Khối cũ
        ws.column_dimensions['G'].width = 12  # Đã dùng
        ws.column_dimensions['H'].width = 12  # Số người
        ws.column_dimensions['I'].width = 15  # Đơn giá
        ws.column_dimensions['J'].width = 18  # Thành tiền
        
        # Title: PHÍ DỊCH VỤ NHÀ : <Tên nhà> - <Địa chỉ>
        title_text = f"PHÍ DỊCH VỤ NHÀ : {house.name} - {house.address}"
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = title_text
        title_cell.font = Font(name='Times New Roman', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40
        
        # Row 2 contains "2" at cell A2 as per sample image
        ws.cell(row=2, column=1, value=2).font = Font(name='Times New Roman', size=10)
        ws.row_dimensions[2].height = 18
        
        # Row 3: Headers
        headers = [
            "Stt", "Tên dịch vụ", "Chỉ số mới", "Chỉ số cũ", 
            "Khối mới", "Khối cũ", "Đã dùng", "Số người", 
            "Đơn giá", "Thành tiền"
        ]
        ws.row_dimensions[3].height = 28
        
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=text)
            cell.font = Font(name='Times New Roman', size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        font_regular = Font(name='Times New Roman', size=11)
        font_bold = Font(name='Times New Roman', size=11, bold=True)
        
        # Row 4 (STT 1): Điện (Số)
        elec_used = max(0, payment.elec_new - payment.elec_old)
        elec_total = elec_used * float(payment.elec_price)
        row_1 = [
            1, "Điện (Số)", 
            payment.elec_new if (payment.elec_new > 0 or payment.elec_old > 0) else "", 
            payment.elec_old if (payment.elec_new > 0 or payment.elec_old > 0) else "",
            "", "", 
            elec_used if (payment.elec_new > 0 or payment.elec_old > 0) else "", 
            "", 
            float(payment.elec_price), 
            elec_total if (payment.elec_new > 0 or payment.elec_old > 0) else float(payment.electricity_fee)
        ]
        
        # Row 5 (STT 2): Nước (Người/tháng)
        water_used = max(0, payment.water_new - payment.water_old)
        water_total = water_used * float(payment.water_price)
        row_2 = [
            2, "Nước (Người/tháng)", "", "",
            payment.water_new if (payment.water_new > 0 or payment.water_old > 0) else "", 
            payment.water_old if (payment.water_new > 0 or payment.water_old > 0) else "",
            water_used if (payment.water_new > 0 or payment.water_old > 0) else "", 
            "", 
            float(payment.water_price), 
            water_total if (payment.water_new > 0 or payment.water_old > 0) else float(payment.water_fee)
        ]
        
        # Row 6 (STT 3): Phí dịch vụ (Người/tháng)
        service_total = payment.service_people * float(payment.service_price)
        row_3 = [
            3, "Phí dịch vụ (Người/tháng)", "", "", "", "", "", 
            payment.service_people, 
            float(payment.service_price), 
            service_total
        ]
        
        # Row 7 (STT 4): Internet (Phòng/tháng)
        row_4 = [
            4, "Internet (Phòng/tháng)", "", "", "", "", "", "", 
            float(payment.internet_price), 
            float(payment.internet_price)
        ]
        
        # Row 8 (STT 5): Tiền phòng
        row_5 = [
            5, "Tiền phòng", 
            room.room_code,  # Yellow room code in Chỉ số mới column
            "", "", "", "", "", "", 
            float(payment.room_fee)
        ]
        
        data_rows = [row_1, row_2, row_3, row_4, row_5]
        if payment.surcharge and float(payment.surcharge) > 0:
            surcharge_label = f"Phụ phí ({payment.surcharge_desc})" if payment.surcharge_desc else "Phụ phí"
            row_6 = [
                6, surcharge_label, "", "", "", "", "", "", "", float(payment.surcharge)
            ]
            data_rows.append(row_6)
        start_row = 4
        
        for i, r_data in enumerate(data_rows):
            current_row_idx = start_row + i
            ws.row_dimensions[current_row_idx].height = 28
            for col_idx, val in enumerate(r_data, 1):
                cell = ws.cell(row=current_row_idx, column=col_idx, value=val)
                cell.font = font_regular
                cell.border = thin_border
                
                # Alignments & formatting
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = font_bold
                elif col_idx == 2:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                elif col_idx in [3, 4, 5, 6, 7, 8]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if val != "":
                        try:
                            # Try to make float/int for display formatting
                            cell.value = float(val) if '.' in str(val) else int(val)
                            cell.number_format = '#,##0'
                        except ValueError:
                            pass
                else:  # Unit Price, Amount columns
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.font = font_bold
                    if val != "":
                        cell.value = float(val)
                        cell.number_format = '#,##0'
                
                # Yellow background for Room Code on Tiền phòng row
                if current_row_idx == 8 and col_idx == 3:
                    cell.fill = yellow_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = font_bold
                    
        # Row 9 (Total Row)
        total_row_idx = start_row + len(data_rows)
        ws.row_dimensions[total_row_idx].height = 28
        
        ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=9)
        total_label_cell = ws.cell(row=total_row_idx, column=1, value="Tổng tiền")
        total_label_cell.font = font_bold
        total_label_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply borders to merged total cells
        for col_idx in range(1, 10):
            ws.cell(row=total_row_idx, column=col_idx).border = thin_border
            
        total_val_cell = ws.cell(row=total_row_idx, column=10, value=float(payment.total_amount))
        total_val_cell.font = Font(name='Times New Roman', size=11, bold=True)
        total_val_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_val_cell.border = thin_border
        total_val_cell.number_format = '#,##0'
        
        # Add bank transfer details at the bottom
        bank_row_idx = total_row_idx + 2
        ws.merge_cells(start_row=bank_row_idx, start_column=1, end_row=bank_row_idx, end_column=10)
        bank_cell = ws.cell(row=bank_row_idx, column=1)
        bank_cell.value = f"số tài khoản : {house.bank_account} Ngân hàng {house.bank_name} Chủ tài khoản: {house.bank_owner}"
        bank_cell.font = Font(name='Times New Roman', size=11, bold=True)
        bank_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[bank_row_idx].height = 20

        note_row_idx = total_row_idx + 3
        ws.merge_cells(start_row=note_row_idx, start_column=1, end_row=note_row_idx, end_column=10)
        note_cell = ws.cell(row=note_row_idx, column=1)
        transfer_code = f"{house.bank_transfer_prefix} {room.room_code}"
        note_cell.value = f"Lưu ý : CHỈ CẦN GHI NỘI DUNG {transfer_code}"
        note_cell.font = Font(name='Times New Roman', size=11, bold=True)
        note_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[note_row_idx].height = 20

        # Build response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Phieu thu phong {room.room_code} - {payment.month}.xlsx"'
        wb.save(response)
        return response


    @action(detail=False, methods=['get'], url_path='export-excel-by-house')
    def export_excel_by_house(self, request):
        """
        Export all payment receipts for a specific house (and optional month) into a single Excel workbook.
        Each payment generates a separate sheet.
        Query params: house_id (required), month (optional, YYYY-MM)
        """
        house_id = request.query_params.get('house_id')
        month = request.query_params.get('month')

        if not house_id:
            return Response(
                {"detail": "Tham số house_id là bắt buộc."},
                status=status.HTTP_400_BAD_REQUEST
            )

        house = get_object_or_404(House, pk=house_id)

        payments_qs = Payment.objects.filter(room__house_id=house_id).select_related('room', 'room__house')
        if month:
            payments_qs = payments_qs.filter(month=month)

        payments_qs = payments_qs.order_by('room__room_code', '-month')

        if not payments_qs.exists():
            return Response(
                {"detail": "Không có phiếu thu nào cho tòa nhà và tháng đã chọn."},
                status=status.HTTP_404_NOT_FOUND
            )

        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        # Remove the default sheet created by openpyxl
        wb.remove(wb.active)

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        font_regular = Font(name='Times New Roman', size=11)
        font_bold = Font(name='Times New Roman', size=11, bold=True)

        for payment in payments_qs:
            room = payment.room

            # Sheet name: max 31 chars (Excel limit)
            sheet_name = f"{room.room_code} - {payment.month}"
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]

            ws = wb.create_sheet(title=sheet_name)
            ws.sheet_view.showGridLines = True

            # Column widths
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 15
            ws.column_dimensions['J'].width = 18

            # Title row
            title_text = f"PHÍ DỊCH VỤ NHÀ : {house.name} - {house.address}"
            ws.merge_cells('A1:J1')
            title_cell = ws['A1']
            title_cell.value = title_text
            title_cell.font = Font(name='Times New Roman', size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 40

            ws.cell(row=2, column=1, value=2).font = Font(name='Times New Roman', size=10)
            ws.row_dimensions[2].height = 18

            # Headers
            headers = [
                "Stt", "Tên dịch vụ", "Chỉ số mới", "Chỉ số cũ",
                "Khối mới", "Khối cũ", "Đã dùng", "Số người",
                "Đơn giá", "Thành tiền"
            ]
            ws.row_dimensions[3].height = 28
            for col_idx, text in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=text)
                cell.font = Font(name='Times New Roman', size=11, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

            # Data rows
            elec_used = max(0, payment.elec_new - payment.elec_old)
            elec_total = elec_used * float(payment.elec_price)
            row_1 = [
                1, "Điện (Số)",
                payment.elec_new if (payment.elec_new > 0 or payment.elec_old > 0) else "",
                payment.elec_old if (payment.elec_new > 0 or payment.elec_old > 0) else "",
                "", "",
                elec_used if (payment.elec_new > 0 or payment.elec_old > 0) else "",
                "",
                float(payment.elec_price),
                elec_total if (payment.elec_new > 0 or payment.elec_old > 0) else float(payment.electricity_fee)
            ]

            water_used = max(0, payment.water_new - payment.water_old)
            water_total = water_used * float(payment.water_price)
            row_2 = [
                2, "Nước (Người/tháng)", "", "",
                payment.water_new if (payment.water_new > 0 or payment.water_old > 0) else "",
                payment.water_old if (payment.water_new > 0 or payment.water_old > 0) else "",
                water_used if (payment.water_new > 0 or payment.water_old > 0) else "",
                "",
                float(payment.water_price),
                water_total if (payment.water_new > 0 or payment.water_old > 0) else float(payment.water_fee)
            ]

            service_total = payment.service_people * float(payment.service_price)
            row_3 = [
                3, "Phí dịch vụ (Người/tháng)", "", "", "", "", "",
                payment.service_people,
                float(payment.service_price),
                service_total
            ]

            row_4 = [
                4, "Internet (Phòng/tháng)", "", "", "", "", "", "",
                float(payment.internet_price),
                float(payment.internet_price)
            ]

            row_5 = [
                5, "Tiền phòng",
                room.room_code,
                "", "", "", "", "", "",
                float(payment.room_fee)
            ]

            data_rows = [row_1, row_2, row_3, row_4, row_5]
            if payment.surcharge and float(payment.surcharge) > 0:
                surcharge_label = f"Phụ phí ({payment.surcharge_desc})" if payment.surcharge_desc else "Phụ phí"
                row_6 = [
                    6, surcharge_label, "", "", "", "", "", "", "", float(payment.surcharge)
                ]
                data_rows.append(row_6)
            start_row = 4

            for i, r_data in enumerate(data_rows):
                current_row_idx = start_row + i
                ws.row_dimensions[current_row_idx].height = 28
                for col_idx, val in enumerate(r_data, 1):
                    cell = ws.cell(row=current_row_idx, column=col_idx, value=val)
                    cell.font = font_regular
                    cell.border = thin_border

                    if col_idx == 1:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = font_bold
                    elif col_idx == 2:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    elif col_idx in [3, 4, 5, 6, 7, 8]:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        if val != "":
                            try:
                                cell.value = float(val) if '.' in str(val) else int(val)
                                cell.number_format = '#,##0'
                            except ValueError:
                                pass
                    else:
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        cell.font = font_bold
                        if val != "":
                            cell.value = float(val)
                            cell.number_format = '#,##0'

                    if current_row_idx == 8 and col_idx == 3:
                        cell.fill = yellow_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = font_bold

            # Total row
            total_row_idx = start_row + len(data_rows)
            ws.row_dimensions[total_row_idx].height = 28
            ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=9)
            total_label_cell = ws.cell(row=total_row_idx, column=1, value="Tổng tiền")
            total_label_cell.font = font_bold
            total_label_cell.alignment = Alignment(horizontal='center', vertical='center')
            for col_idx in range(1, 10):
                ws.cell(row=total_row_idx, column=col_idx).border = thin_border
            total_val_cell = ws.cell(row=total_row_idx, column=10, value=float(payment.total_amount))
            total_val_cell.font = Font(name='Times New Roman', size=11, bold=True)
            total_val_cell.alignment = Alignment(horizontal='right', vertical='center')
            total_val_cell.border = thin_border
            total_val_cell.number_format = '#,##0'

            # Add bank transfer details at the bottom
            bank_row_idx = total_row_idx + 2
            ws.merge_cells(start_row=bank_row_idx, start_column=1, end_row=bank_row_idx, end_column=10)
            bank_cell = ws.cell(row=bank_row_idx, column=1)
            bank_cell.value = f"số tài khoản : {house.bank_account} Ngân hàng {house.bank_name} Chủ tài khoản: {house.bank_owner}"
            bank_cell.font = Font(name='Times New Roman', size=11, bold=True)
            bank_cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[bank_row_idx].height = 20

            note_row_idx = total_row_idx + 3
            ws.merge_cells(start_row=note_row_idx, start_column=1, end_row=note_row_idx, end_column=10)
            note_cell = ws.cell(row=note_row_idx, column=1)
            transfer_code = f"{house.bank_transfer_prefix} {room.room_code}"
            note_cell.value = f"Lưu ý : CHỈ CẦN GHI NỘI DUNG {transfer_code}"
            note_cell.font = Font(name='Times New Roman', size=11, bold=True)
            note_cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[note_row_idx].height = 20

        # Build filename
        month_label = f" - {month}" if month else ""
        safe_house_name = house.name.replace(' ', '_')
        filename = f"Phieu_thu_{safe_house_name}{month_label}.xlsx"

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class ExpenseViewSet(viewsets.ModelViewSet):
    """
    API endpoint for Expenses (Chi phí)
    """
    queryset = Expense.objects.all()
    serializer_class = ExpenseSerializer
    pagination_class = OptionalPageNumberPagination

    def get_queryset(self):
        queryset = Expense.objects.all()
        
        # Filter by house
        house_id = self.request.query_params.get('house_id', None)
        if house_id:
            queryset = queryset.filter(house_id=house_id)
            
        # Filter by category
        category = self.request.query_params.get('category', None)
        if category:
            queryset = queryset.filter(category=category)
            
        return queryset


class DashboardStatsView(APIView):
    """
    API endpoint to retrieve general statistics for the dashboard
    """
    def get(self, request, format=None):
        # House counts
        total_houses = House.objects.count()
        
        # Room counts
        total_rooms = Room.objects.count()
        rented_rooms = Room.objects.filter(status='rented').count()
        empty_rooms = Room.objects.filter(status='empty').count()
        
        # Tenant count
        total_tenants = Tenant.objects.count()
        
        # Current month revenue and expenses
        today = timezone.now().date()
        current_month = today.strftime('%Y-%m')
        current_year = today.year
        current_month_num = today.month
        
        current_month_revenue_agg = Payment.objects.filter(
            month=current_month, 
            payment_status='paid'
        ).aggregate(total=Sum('total_amount'))
        current_month_revenue = current_month_revenue_agg['total'] or 0
        
        current_month_expense_agg = Expense.objects.filter(
            date__year=current_year,
            date__month=current_month_num
        ).aggregate(total=Sum('amount'))
        current_month_expense = current_month_expense_agg['total'] or 0
        
        current_month_net = current_month_revenue - current_month_expense
        
        # 1. Revenue & Expense chart data (Last 6 months)
        revenue_by_month = []
        
        for i in range(5, -1, -1):
            m = current_month_num - i
            y = current_year
            if m <= 0:
                m += 12
                y -= 1
            
            month_str = f"{y:04d}-{m:02d}"
            month_label = f"Tháng {m}/{y}"
            
            rev = Payment.objects.filter(
                month=month_str, 
                payment_status='paid'
            ).aggregate(total=Sum('total_amount'))['total'] or 0
            
            exp = Expense.objects.filter(
                date__year=y,
                date__month=m
            ).aggregate(total=Sum('amount'))['total'] or 0
            
            revenue_by_month.append({
                'month': month_str,
                'label': month_label,
                'revenue': float(rev),
                'expense': float(exp),
                'profit': float(rev) - float(exp)
            })
            
        # 2. Payment status count in current month
        unpaid_count = Payment.objects.filter(month=current_month, payment_status='unpaid').count()
        paid_count = Payment.objects.filter(month=current_month, payment_status='paid').count()
        
        # 3. Recent unpaid invoices
        recent_unpaid = PaymentSerializer(
            Payment.objects.filter(payment_status='unpaid').order_by('-created_at')[:5],
            many=True
        ).data
        
        return Response({
            'stats': {
                'total_houses': total_houses,
                'total_rooms': total_rooms,
                'rented_rooms': rented_rooms,
                'empty_rooms': empty_rooms,
                'total_tenants': total_tenants,
                'current_month_revenue': float(current_month_revenue),
                'current_month_expense': float(current_month_expense),
                'current_month_net': float(current_month_net),
                'current_month': current_month,
            },
            'revenue_by_month': revenue_by_month,
            'payment_status_distribution': [
                { 'name': 'Đã thanh toán', 'value': paid_count, 'color': '#10B981' },
                { 'name': 'Chưa thanh toán', 'value': unpaid_count, 'color': '#EF4444' }
            ],
            'recent_unpaid': recent_unpaid
        }, status=status.HTTP_200_OK)


class LogoutView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        request.user.auth_token.delete()
        return Response({"detail": "Đăng xuất thành công."}, status=status.HTTP_200_OK)
